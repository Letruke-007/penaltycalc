from __future__ import annotations

from typing import Any, Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from .styles import FONT_NAME, MONEY_FMT, THIN, THICK


# Excel often stores an explicit (empty) diagonal side; keep it to match model XLSX 1:1
EMPTY_DIAG = Side(style=None, color=None)


# ---------------------------------------------------------------------
# Layout helpers
# ---------------------------------------------------------------------

def apply_column_widths(ws: Worksheet, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def apply_row_heights(ws: Worksheet, heights: dict[int, float]) -> None:
    for r, h in heights.items():
        ws.row_dimensions[r].height = h


# ---------------------------------------------------------------------
# Cell setters (ЕДИНАЯ ТОЧКА СТИЛЕЙ)
# ---------------------------------------------------------------------

def set_cell(
    ws: Worksheet,
    addr: str,
    value: Any,
    *,
    bold: bool = False,
    size: int = 11,
    h: str = "center",
    v: str = "center",
    fill: Optional[PatternFill] = None,
    num_fmt: Optional[str] = None,
    wrap: Optional[bool] = None,   # перенос по умолчанию как в эталоне (None)
    italic: bool = False,
) -> None:
    cell = ws[addr]
    cell.value = value
    cell.alignment = Alignment(
        horizontal=h,
        vertical=v,
        wrap_text=wrap,
    )
    cell.font = Font(
        name=FONT_NAME,
        size=size,
        bold=bold,
        italic=italic,
    )
    if fill is not None:
        cell.fill = fill
    if num_fmt is not None:
        cell.number_format = num_fmt


def set_money(
    ws: Worksheet,
    addr: str,
    amount: float,
    *,
    size: int = 10,
    bold: bool = False,
    fill: Optional[PatternFill] = None,
) -> None:
    cell = ws[addr]
    cell.value = amount
    cell.number_format = MONEY_FMT
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=False,  # <<< числа НИКОГДА не переносятся
    )
    cell.font = Font(
        name=FONT_NAME,
        size=size,
        bold=bold,
    )
    if fill is not None:
        cell.fill = fill


# ---------------------------------------------------------------------
# Borders
# ---------------------------------------------------------------------

def apply_thick_outer_border(ws: Worksheet, top_left: str, bottom_right: str) -> None:
    tl = ws[top_left]
    br = ws[bottom_right]
    min_row, min_col = tl.row, tl.column
    max_row, max_col = br.row, br.column

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = Border(
                left=THICK if c == min_col else THIN,
                right=THICK if c == max_col else THIN,
                top=THICK if r == min_row else THIN,
                bottom=THICK if r == max_row else THIN,
               diagonal=EMPTY_DIAG,
            )


def apply_thin_outer_border(ws: Worksheet, top_left: str, bottom_right: str) -> None:
    tl = ws[top_left]
    br = ws[bottom_right]
    min_row, min_col = tl.row, tl.column
    max_row, max_col = br.row, br.column

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = Border(
                left=THIN if c == min_col else cell.border.left,
                right=THIN if c == max_col else cell.border.right,
                top=THIN if r == min_row else cell.border.top,
                bottom=THIN if r == max_row else cell.border.bottom,
               diagonal=EMPTY_DIAG,
            )


def apply_thin_grid(ws: Worksheet, top_left: str, bottom_right: str) -> None:
    tl = ws[top_left]
    br = ws[bottom_right]
    min_row, min_col = tl.row, tl.column
    max_row, max_col = br.row, br.column

    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN, diagonal=EMPTY_DIAG)

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = border
