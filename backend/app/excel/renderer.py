# backend/app/excel/renderer.py
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from typing import Any, Optional, Tuple, List

from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins

from .footnotes import rate_share_footnote
from ..contracts.statement import Statement
from .calc_rows import CalcRow, build_calc_rows
from .styles import (
    COLUMN_WIDTHS,
    DATE_FMT,
    FILL_GRAY,
    FILL_YELLOW,
    FILL_YELLOW_LIGHT,
    ROW_HEIGHTS,
    FILL_TOTAL_PEACH,
    FILL_TOTAL_YELLOW,
    FILL_TOTAL_LIME,
    FILL_TOTAL_BLUE,
    MONEY_FMT,
)

# FILL_GREEN может отсутствовать в старой версии styles.py — делаем fallback
try:
    from .styles import FILL_GREEN  # type: ignore
except Exception:
    FILL_GREEN = PatternFill("solid", fgColor="FFDDE8CB")  # fallback как в эталоне

from .style_apply import (
    apply_column_widths,
    apply_thick_outer_border,
    apply_thin_outer_border,
    apply_row_heights,
    apply_thin_grid,
    set_cell,
    set_money,
)

# =============================================================================
# Print setup helpers
# =============================================================================

def _find_last_nonempty_row(ws: Worksheet, *, min_col: int = 1, max_col: int = 13) -> int:
    """
    Returns the last row index that has at least one non-empty cell value
    in [min_col..max_col]. Empty means: None or "" or whitespace-only string.
    """
    max_r = ws.max_row or 1
    for r in range(max_r, 0, -1):
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            return r
    return 1


def _apply_print_setup(ws: Worksheet) -> None:
    """
    Applies page setup for printing:
      - Print area: A1:M(last_nonempty_row+1)
      - Orientation: landscape
      - Paper: A4
      - Fit to width: 1 page
      - Fit to height: auto (0 / unlimited)
      - Margins: ~0.5–0.7 cm (use 0.25 inch ≈ 0.635 cm)
    """
    last_row = _find_last_nonempty_row(ws, min_col=1, max_col=13)
    print_last = last_row + 1
    ws.print_area = f"A1:M{print_last}"

    # Page setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Fit to one page wide, unlimited height
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 == auto / no limit

    # Margins are in inches in openpyxl
    m = 0.25  # ~0.635 cm
    ws.page_margins = PageMargins(
        left=m,
        right=m,
        top=m,
        bottom=m,
        header=0.2,
        footer=0.2,
    )


# =============================================================================
# Footer helpers
# =============================================================================

def _merge_footnote_block_a_h(ws: Worksheet, start_row: int) -> None:
    """Merge the footnote text cell area to match the model XLSX.

    Required merge area:
      - A..H in the footnote row (start_row)
      - A..H in the next two rows (start_row+1, start_row+2)

    Equivalent merged range: A{start_row}:H{start_row+2}.

    Alignment:
      - vertical: center (within the 3-row merged region)
      - horizontal: left
    """
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 2, end_column=8)

    cell = ws.cell(row=start_row, column=1)
    cell.alignment = cell.alignment.copy(horizontal="left", vertical="center", wrap_text=True)


# =============================================================================
# Public API
# =============================================================================

def render_statement_sheet(ws: Worksheet, stmt: Statement) -> None:
    """
    Render:
      - Header A2:M14 (exact formatting)
      - Rows from 15: ledger + penalty (G–M) + subtotals + totals + footer

    IMPORTANT:
      Rendering is fully here; pipeline must not style anything.
    """
    body = stmt.statement

    # Freeze panes disabled (visible separator line between 14 and 15 in Excel UI)
    ws.freeze_panes = None

    # widths + fixed header row heights
    apply_column_widths(ws, COLUMN_WIDTHS)
    apply_row_heights(ws, ROW_HEIGHTS)

    _render_header(ws, body)

    # Точечная правка рамок (как в эталоне): у B8:C8 и B10:C10 правая граница должна быть thick
    from openpyxl.styles import Side
    thick = Side(style="thick", color="FF000000")
    for addr in ("B8", "C8", "B10", "C10"):
        cell = ws[addr]
        cell.border = cell.border.copy(right=thick)

    rows, _params = build_calc_rows(stmt)

    key_rate = float(body.rate_percent) / 100.0

    tail = _render_calc_rows_g_h_k_i(
        ws,
        start_row=15,
        rows=rows,
        key_rate=key_rate,
        contract_number=body.contract.number,
        category=body.category or "",
    )

    # Apply thin grid for whole table (header + data + totals + footer if included)
    if tail.last_row >= 12:
        apply_thin_grid(ws, "A12", f"M{tail.total_row or tail.last_row}")

    # Set dynamic row heights for footer and "СПРАВОЧНО" block to match the look
    dyn_heights: dict[int, float] = {}
    if tail.footer_start_row is not None:
        # 3 footer rows (footnote block), plus "СПРАВОЧНО" and 3 rows beneath
        dyn_heights[tail.footer_start_row] = 16.9
        dyn_heights[tail.footer_start_row + 1] = 16.9
        dyn_heights[tail.footer_start_row + 2] = 16.9
        if tail.sprav_start_row is not None:
            dyn_heights[tail.sprav_start_row] = 16.9
            dyn_heights[tail.sprav_start_row + 1] = 16.9
            dyn_heights[tail.sprav_start_row + 2] = 16.5
            dyn_heights[tail.sprav_start_row + 3] = 16.5

    if dyn_heights:
        apply_row_heights(ws, dyn_heights)

    # ✅ Apply print settings at the very end (sheet fully populated)
    _apply_print_setup(ws)


def render_statements_sheet(ws: Worksheet, stmts: List[Statement]) -> None:
    if not stmts:
        return

    def _sort_key(s: Statement) -> tuple:
        body = s.statement
        cn = (getattr(body.contract, "number", "") or "")
        cd = (getattr(body.contract, "date", "") or "")
        src = (getattr(body, "source_pdf", "") or "")
        return (str(cn), str(cd), str(src))

    stmts_sorted = sorted(stmts, key=_sort_key)

    ws.freeze_panes = None
    apply_column_widths(ws, COLUMN_WIDTHS)

    first_body = stmts_sorted[0].statement
    _render_title_and_debtor(ws, first_body)

    contract_footer_debt_cells: list[str] = []
    contract_footer_penalty_cells: list[str] = []

    cursor_row = 6

    for stmt in stmts_sorted:
        body = stmt.statement
        data_start_row = _render_contract_header_at(ws, body, start_row=cursor_row)

        rows, _params = build_calc_rows(stmt)
        key_rate = float(body.rate_percent) / 100.0

        tail = _render_calc_rows_g_h_k_i(
            ws,
            start_row=data_start_row,
            rows=rows,
            key_rate=key_rate,
            contract_number=body.contract.number,
            category=body.category or "",
            include_sprav=False,
        )

        if tail.footer_start_row is not None:
            contract_footer_debt_cells.append(f"M{tail.footer_start_row}")
            contract_footer_penalty_cells.append(f"M{tail.footer_start_row + 1}")

        # применяем grid ТОЛЬКО к табличной части договора
        table_top = data_start_row - 3   # строка заголовков (Период / Примечание / ...)
        table_bottom = (
            tail.total_row
            if tail.total_row is not None
            else tail.last_row
        )

        if table_bottom >= table_top:
            apply_thin_grid(ws, f"A{table_top}", f"M{table_bottom}")

        # dynamic row heights for each footer block
        dyn: dict[int, float] = {}
        if tail.footer_start_row is not None:
            dyn[tail.footer_start_row] = 16.9
            dyn[tail.footer_start_row + 1] = 16.9
            dyn[tail.footer_start_row + 2] = 16.9
        if dyn:
            apply_row_heights(ws, dyn)

        cursor_row = tail.last_row + 3  # +2 пустых строки

    if contract_footer_debt_cells and contract_footer_penalty_cells:
        sprav_start = cursor_row
        _render_sprav_block_sum(ws, sprav_start, contract_footer_debt_cells, contract_footer_penalty_cells)

        # В СПРАВОЧНО не применяем grid/outer borders (как в single)
        apply_row_heights(ws, {
            sprav_start: 16.9,
            sprav_start + 1: 16.9,
            sprav_start + 2: 16.5,
            sprav_start + 3: 16.5,
        })

    # ✅ Apply print settings at the very end (sheet fully populated)
    _apply_print_setup(ws)


# =============================================================================
# Multi render helpers (for merged XLSX)
# =============================================================================

def _render_title_and_debtor(ws: Worksheet, body: Any) -> None:
    """Render only the common top part (rows 2..5): title + debtor."""
    debtor_name = (body.debtor.name or "").upper()

    set_cell(
        ws,
        "A2",
        "Расчет долга и неустойки в связи с просрочкой в погашении задолженности за поставленные энергоресурсы",
        bold=True,
        size=16,
        h="left",
        v="center",
    )

    set_cell(ws, "A4", "Должник:", bold=True, size=13, h="left", v="center")
    ws.merge_cells("B4:M4")
    set_cell(ws, "B4", debtor_name, bold=True, size=13, h="left", v="center")

    # In single-contract template these heights are critical; keep them here too.
    apply_row_heights(ws, {2: 20.65, 4: 16.9})


def _render_contract_header_at(ws: Worksheet, body: Any, *, start_row: int) -> int:
    """Render contract header + table header with a vertical offset.

    Returns the first data row (equivalent to row 15 in the single-contract template).
    """
    contract_no = body.contract.number
    contract_date = body.contract.date
    category = body.category or ""
    calc_date = body.calc_date

    r_osn = start_row
    r_cat = start_row + 2
    r_calc = start_row + 4
    r_h12 = start_row + 6
    r_h13 = start_row + 7
    r_nums = start_row + 8
    r_data = start_row + 9

    # Основание
    set_cell(ws, f"A{r_osn}", "Основание:", bold=True, size=13, h="left", v="center")
    ws.merge_cells(f"B{r_osn}:M{r_osn}")
    set_cell(
        ws,
        f"B{r_osn}",
        f"Договор № {contract_no} от {contract_date}",
        bold=True,
        size=13,
        h="left",
        v="center",
        fill=FILL_GRAY,
    )

    # Категория (B..C merged + thick outer border)
    set_cell(ws, f"A{r_cat}", "Категория:", bold=True, size=13, h="left", v="center")
    ws.merge_cells(f"B{r_cat}:C{r_cat}")

    DISPLAY_CATEGORY = {
        "Управляющая организация": "УК",
        "ТСЖ, ЖСК, ЖК": "ТСЖ, ЖСК, ЖК",
        "Прочие": "Прочие",
        "Собственники жилых помещений в МКД": "Собственники жилых помещений в МКД",
        "Собственники нежилых помещений в МКД": "Собственники нежилых помещений в МКД",
        "УК": "УК",
    }
    category_display = DISPLAY_CATEGORY.get(category, category)
    set_cell(ws, f"B{r_cat}", category_display, bold=True, size=13, h="center", v="center")
    apply_thick_outer_border(ws, f"B{r_cat}", f"C{r_cat}")

    # На дату (B..C merged + thick outer border)
    set_cell(ws, f"A{r_calc}", "На дату:", bold=True, size=13, h="left", v="center")
    ws.merge_cells(f"B{r_calc}:C{r_calc}")
    dt_calc = _parse_date_ddmmyyyy(calc_date)
    set_cell(ws, f"B{r_calc}", dt_calc, bold=True, size=13, h="center", v="center", num_fmt=DATE_FMT)
    apply_thick_outer_border(ws, f"B{r_calc}", f"C{r_calc}")

    # Rows r_h12–r_h13: column headers (same as in _render_header)
    headers_af = [
        ("A", "Период"),
        ("B", "Примечание"),
        ("C", "Начислено"),
        ("D", "Оплачено"),
        ("E", "Дата"),
        ("F", "Долг"),
    ]
    for col, title in headers_af:
        ws.merge_cells(f"{col}{r_h12}:{col}{r_h13}")
        set_cell(ws, f"{col}{r_h12}", title, bold=True, size=11, h="center", v="center", fill=FILL_YELLOW)

    ws.merge_cells(f"G{r_h12}:I{r_h12}")
    set_cell(ws, f"G{r_h12}", "Период просрочки", bold=True, size=11, h="left", v="center", fill=FILL_YELLOW)
    set_cell(ws, f"G{r_h13}", "с", size=11, h="center", v="center", fill=FILL_YELLOW_LIGHT)
    set_cell(ws, f"H{r_h13}", "по", size=11, h="center", v="center", fill=FILL_YELLOW_LIGHT)
    set_cell(ws, f"I{r_h13}", "дней", size=11, h="center", v="center", fill=FILL_YELLOW_LIGHT)

    headers_jm = [("J", "Ставка"), ("K", "Доля ставки*"), ("L", "Формула"), ("M", "Неустойка")]
    for col, title in headers_jm:
        ws.merge_cells(f"{col}{r_h12}:{col}{r_h13}")
        set_cell(ws, f"{col}{r_h12}", title, bold=True, size=11, h="center", v="center", fill=FILL_YELLOW)

    for col in "ABCDEF":
        apply_thin_outer_border(ws, f"{col}{r_h12}", f"{col}{r_h13}")
    apply_thin_outer_border(ws, f"G{r_h12}", f"I{r_h12}")
    for col in "GHI":
        apply_thin_outer_border(ws, f"{col}{r_h13}", f"{col}{r_h13}")
    for col in "JKLM":
        apply_thin_outer_border(ws, f"{col}{r_h12}", f"{col}{r_h13}")

    # Row r_nums: column numbers 1..13
    for idx, col in enumerate(list("ABCDEFGHIJKLM"), start=1):
        set_cell(ws, f"{col}{r_nums}", idx, size=11, h="center", v="center", fill=FILL_GRAY)
        apply_thin_outer_border(ws, f"{col}{r_nums}", f"{col}{r_nums}")

    # match template heights for these header rows
    apply_row_heights(ws, {r_osn: 16.9, r_cat: 16.9, r_calc: 16.9})

    # same thick-right border fix as in render_statement_sheet for merged header cells
    from openpyxl.styles import Side
    thick = Side(style="thick", color="FF000000")
    for addr in (f"B{r_cat}", f"C{r_cat}", f"B{r_calc}", f"C{r_calc}"):
        cell = ws[addr]
        cell.border = cell.border.copy(right=thick)

    return r_data


def _render_sprav_block_sum(ws: Worksheet, sprav_start: int, debt_cells: list[str], penalty_cells: list[str]) -> None:
    """Global СПРАВОЧНО block for merged XLSX (sums across contracts).
    Как в single: без сетки/рамок, только тонкая нижняя линия под заголовком 'СПРАВОЧНО' (J:M).
    """
    set_cell(ws, f"J{sprav_start}", "СПРАВОЧНО", bold=True, size=13, h="left", v="center")

    # Bottom border for "СПРАВОЧНО" row (J..M)
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="FF000000")
    for col in ("J", "K", "L", "M"):
        c = ws[f"{col}{sprav_start}"]
        c.border = (c.border.copy(bottom=thin) if c.border else Border(bottom=thin))

    debt_sum = "=SUM(" + ",".join(debt_cells) + ")" if debt_cells else "=0"
    pen_sum = "=SUM(" + ",".join(penalty_cells) + ")" if penalty_cells else "=0"

    set_cell(ws, f"J{sprav_start+1}", "Цена иска, руб.", size=13, h="left", v="center")
    set_cell(
        ws,
        f"M{sprav_start+1}",
        f"=M{sprav_start+2}+M{sprav_start+3}",
        bold=True,
        size=13,
        h="right",
        v="center",
        num_fmt=MONEY_FMT,
        fill=FILL_TOTAL_BLUE,
    )

    set_cell(ws, f"J{sprav_start+2}", "в т.ч. основной долг", size=13, h="left", v="center")
    set_cell(ws, f"M{sprav_start+2}", debt_sum, size=13, h="right", v="center", num_fmt=MONEY_FMT)

    set_cell(ws, f"J{sprav_start+3}", "в т.ч. неустойка", size=13, h="left", v="center")
    set_cell(ws, f"M{sprav_start+3}", pen_sum, size=13, h="right", v="center", num_fmt=MONEY_FMT)


# =============================================================================
# Calc rows rendering (G–M inside row blocks)
# =============================================================================

@dataclass(frozen=True)
class RenderTail:
    last_row: int
    total_row: Optional[int]
    footer_start_row: Optional[int]
    sprav_start_row: Optional[int]

import math

def _autofit_wrapped_row_height_for_col(
    ws: Worksheet,
    *,
    row: int,
    col_letter: str,
    text: str,
    font_size: int = 10,
    base_row_height: float = 16.9,
) -> None:
    """
    Deterministic "auto-fit" for wrapped text.
    Excel won't auto-adjust row height on open; we approximate by column width.
    Only increases height (never decreases).
    """
    if not text:
        return

    # Column width from COLUMN_WIDTHS; fallback to current sheet width or a safe default.
    col_w = COLUMN_WIDTHS.get(col_letter)
    if not col_w:
        try:
            col_w = float(ws.column_dimensions[col_letter].width or 20.0)
        except Exception:
            col_w = 20.0

    # Rough chars per line for Arial 10/11 in Excel:
    # ~1.1–1.25 chars per width unit depending on font/zoom. Use conservative.
    chars_per_line = max(12, int(col_w * 1.15))

    # Count visual lines (respect explicit newlines too)
    parts = str(text).splitlines() or [str(text)]
    lines = 0
    for p in parts:
        p = p.strip()
        if not p:
            lines += 1
        else:
            lines += max(1, math.ceil(len(p) / chars_per_line))

    # Make row taller. For wrapped cells inside merged A-range, total height = sum of row heights,
    # so lifting the FIRST row of the merged block is enough.
    target = max(base_row_height, base_row_height * lines)

    cur = ws.row_dimensions[row].height
    if cur is None or cur < target:
        ws.row_dimensions[row].height = target


def _render_calc_rows_g_h_k_i(
    ws: Worksheet,
    *,
    start_row: int,
    rows: list[CalcRow],
    key_rate: float,
    contract_number: str,
    category: str,
    include_sprav: bool = True,
) -> RenderTail:

    """
    Renders:
      - blocks (charge row starts a block)
      - merges A/B/C inside each block
      - "ИТОГО ПО ПЕРИОДУ" after each block
      - "ИТОГО" after all blocks
      - footer with footnote + totals + sprav block (single mode only)
    """
    r = start_row
    block_start: Optional[int] = None
    block_label_by_start: dict[int, str] = {}
    subtotal_rows: list[int] = []
    
    AA_PREFIX = "Доля от размера годовой корректировки"

    def _estimate_wrapped_lines_for_col(text: str, col_letter: str) -> int:
        if not text:
            return 1
        col_w = COLUMN_WIDTHS.get(col_letter) or float(ws.column_dimensions[col_letter].width or 20.0)
        # conservative chars per line
        chars_per_line = max(12, int(col_w * 1.15))
        parts = str(text).splitlines() or [str(text)]
        lines = 0
        for p in parts:
            p = p.strip()
            lines += max(1, math.ceil(len(p) / chars_per_line)) if p else 1
        return max(1, lines)

    def _distribute_height_over_block(block_start_row: int, block_end_row: int, text: str) -> None:
        # We want TOTAL height sufficient for wrapped text in merged cell A{start}:A{end}
        base_h = ws.row_dimensions[block_start_row].height or 16.9  # default-like
        lines_needed = _estimate_wrapped_lines_for_col(text, "A")

        # Total target height for the merged cell
        total_target = max(base_h, base_h * lines_needed)

        nrows = max(1, block_end_row - block_start_row + 1)
        per_row = total_target / nrows

        # Cap per-row height to avoid "fat" blocks even if text is very long.
        per_row = min(max(base_h, per_row), 26.0)

        for rr in range(block_start_row, block_end_row + 1):
            cur = ws.row_dimensions[rr].height
            if cur is None or cur < per_row:
                ws.row_dimensions[rr].height = per_row


    def close_block_and_subtotal(block_start_row: int, block_end_row: int) -> int:
        # merge A/B/C over the block
        ws.merge_cells(f"A{block_start_row}:A{block_end_row}")
        ws.merge_cells(f"B{block_start_row}:B{block_end_row}")
        ws.merge_cells(f"C{block_start_row}:C{block_end_row}")

        
        lbl = block_label_by_start.get(block_start_row, "") or ""
        if lbl.strip().startswith(AA_PREFIX):
            # Ensure wrap is enabled on the merged top-left cell (safety)
            ws[f"A{block_start_row}"].alignment = ws[f"A{block_start_row}"].alignment.copy(wrap_text=True, vertical="center")
            _distribute_height_over_block(block_start_row, block_end_row, lbl)

        # subtotal row right after block
        sr = block_end_row + 1

        set_cell(ws, f"A{sr}", "", bold=True, size=10, h="center", v="center")
        set_cell(ws, f"B{sr}", "ИТОГО ПО ПЕРИОДУ", bold=True, size=10, h="right", v="center")

        set_cell(ws, f"C{sr}", f"=SUM(C{block_start_row}:C{block_end_row})", bold=True, size=10, h="center", v="center", num_fmt=MONEY_FMT)
        set_cell(ws, f"D{sr}", f"=SUM(D{block_start_row}:D{block_end_row})", bold=True, size=10, h="center", v="center", num_fmt=MONEY_FMT)
        set_cell(ws, f"F{sr}", f"=C{sr}-D{sr}", bold=True, size=10, h="center", v="center", num_fmt=MONEY_FMT)
        set_cell(ws, f"M{sr}", f"=SUM(M{block_start_row}:M{block_end_row})", bold=True, size=10, h="center", v="center", num_fmt=MONEY_FMT)

        # Остальные ячейки строки подытога тоже bold в эталоне
        for col in "EGHIJKL":
            set_cell(ws, f"{col}{sr}", None if col != "L" else "", bold=True, size=10, h="center", v="center")

        # Fill subtotal row: full-row gray
        for col in "ABCDEFGHIJKLM":
            ws[f"{col}{sr}"].fill = FILL_GRAY

        subtotal_rows.append(sr)
        return sr

    total_row: Optional[int] = None
    footer_start: Optional[int] = None
    sprav_start: Optional[int] = None

    for row in rows:
        # Charge row starts a new block:
        # In our CalcRow model: charge row has period_label and charged != None
        is_charge_row = bool(row.period_label) and (row.charged is not None)

        if is_charge_row:
            # close previous block
            if block_start is not None:
                block_end = r - 1
                r = close_block_and_subtotal(block_start, block_end) + 1
            block_start = r

            is_aa_block = bool(row.period_label) and row.period_label.strip().startswith(AA_PREFIX)

            # A: for AA we enable wrap, BUT keep vertical center (so ordinary months stay centered)
            set_cell(ws, f"A{r}", row.period_label, size=10, h="center", v="center", wrap=is_aa_block)

            # remember label for the block (used on close)
            block_label_by_start[r] = row.period_label


            # Auto-height only for long labels (AA blocks)
            if row.period_label and len(row.period_label) > 25:
                _autofit_wrapped_row_height_for_col(
                    ws,
                    row=r,
                    col_letter="A",
                    text=row.period_label,
                    font_size=10,
                    base_row_height=16.9,
             )

            set_cell(ws, f"B{r}", row.note or "-", size=10, h="center", v="center")
            set_money(ws, f"C{r}", float(row.charged), size=10)
        else:
            # inside block A/B/C will be merged => keep empty cells
            set_cell(ws, f"A{r}", None, size=10, h="center", v="center")
            set_cell(ws, f"B{r}", None, size=10, h="center", v="center")
            set_cell(ws, f"C{r}", None, size=10, h="center", v="center")

        # D/E: payment
        if row.paid is not None:
            set_money(ws, f"D{r}", float(row.paid), size=10)
            if row.pay_date is not None:
                set_cell(ws, f"E{r}", row.pay_date, size=10, h="center", v="center", num_fmt=DATE_FMT)
            else:
                set_cell(ws, f"E{r}", "-", size=10, h="center", v="center")
        else:
            set_money(ws, f"D{r}", 0.0, size=10)
            set_cell(ws, f"E{r}", "-", size=10, h="center", v="center")

        # F: debt formula (exact template logic)
        if is_charge_row:
            set_cell(ws, f"F{r}", f"=C{r}-D{r}", size=10, h="center", v="center", num_fmt=MONEY_FMT)
        else:
            set_cell(ws, f"F{r}", f"=F{r-1}-D{r}", size=10, h="center", v="center", num_fmt=MONEY_FMT)

        # G/H: overdue dates
        if row.overdue_from is not None:
            set_cell(ws, f"G{r}", row.overdue_from, size=10, h="center", v="center", num_fmt=DATE_FMT)
        else:
            set_cell(ws, f"G{r}", None, size=10, h="center", v="center")

        if row.overdue_to is not None:
            set_cell(ws, f"H{r}", row.overdue_to, size=10, h="center", v="center", num_fmt=DATE_FMT)
        else:
            set_cell(ws, f"H{r}", None, size=10, h="center", v="center")

        # I: days formula
        if row.overdue_from is not None and row.overdue_to is not None:
            set_cell(ws, f"I{r}", f"=H{r}-G{r}+1", size=10, h="center", v="center", num_fmt="0")
        else:
            set_cell(ws, f"I{r}", None, size=10, h="center", v="center")

        # J: key rate (only when overdue exists)
        has_overdue = bool(row.overdue_from and row.overdue_to)
        set_cell(
            ws,
            f"J{r}",
            key_rate if has_overdue else None,
            size=10,
            h="center",
            v="center",
            num_fmt="0.00%",
        )

        # K: fraction
        if row.fraction is not None and has_overdue:
            set_cell(ws, f"K{r}", float(row.fraction), size=10, h="center", v="center", num_fmt=r"#\ ??/???")
        else:
            set_cell(ws, f"K{r}", None, size=10, h="center", v="center")

        # L: formula text (from calc_rows)
        set_cell(ws, f"L{r}", row.formula_text or "", size=10, h="center", v="center")

        # M: penalty
        has_penalty = bool(has_overdue and row.fraction is not None)
        if has_penalty:
            set_cell(ws, f"M{r}", f"=F{r}*I{r}*J{r}*K{r}", size=10, h="center", v="center", num_fmt=MONEY_FMT)
        else:
            set_cell(ws, f"M{r}", None, size=10, h="center", v="center")

        r += 1

    # close last block if any
    if block_start is not None and r > block_start:
        block_end = r - 1
        r = close_block_and_subtotal(block_start, block_end) + 1

    # Grand total + footer
    if subtotal_rows:
        total_row = r

        set_cell(ws, f"A{total_row}", "", bold=True, size=10, h="center", v="center")
        set_cell(ws, f"B{total_row}", "ИТОГО", bold=True, size=10, h="center", v="center")

        c_args = ",".join(f"C{sr}" for sr in subtotal_rows)
        d_args = ",".join(f"D{sr}" for sr in subtotal_rows)
        m_args = ",".join(f"M{sr}" for sr in subtotal_rows)

        set_cell(ws, f"C{total_row}", f"=SUM({c_args})", bold=True, size=10, h="center", v="center", num_fmt=MONEY_FMT)
        set_cell(ws, f"D{total_row}", f"=SUM({d_args})", bold=True, size=10, h="center", v="center", num_fmt=MONEY_FMT)
        set_cell(ws, f"F{total_row}", f"=C{total_row}-D{total_row}", bold=True, size=10, h="center", v="center", num_fmt=MONEY_FMT)
        set_cell(ws, f"M{total_row}", f"=SUM({m_args})", bold=True, size=10, h="center", v="center", num_fmt=MONEY_FMT)

        for col in "EGHIJKL":
            set_cell(ws, f"{col}{total_row}", None if col != "L" else "", bold=True, size=10, h="center", v="center")

        for col in "ABCDEFGHIJKLM":
            ws[f"{col}{total_row}"].fill = FILL_TOTAL_PEACH

        r = total_row + 1

        # Footer (как в single) — начинается через 1 пустую строку
        r += 1
        footer_start = r

        footnote = rate_share_footnote(contract_number, category)

        set_cell(ws, f"A{footer_start}", footnote, size=9, italic=True, h="left", v="center", wrap=True)
        
        # For star-prefixed legal footnotes ("* доля ставки ..." and other variants)
        # the template uses a 3-row merged area A..H.
        if isinstance(footnote, str) and footnote.strip().startswith("*"):
            _merge_footnote_block_a_h(ws, footer_start)

        set_cell(ws, f"J{footer_start}", "Всего долг, руб.", size=10, h="left", v="center")
        set_cell(ws, f"M{footer_start}", f"=F{total_row}", size=10, h="center", v="center", num_fmt=MONEY_FMT, fill=FILL_GREEN)

        set_cell(ws, f"J{footer_start+1}", "Всего неустойка, руб.", size=10, h="left", v="center")
        set_cell(ws, f"M{footer_start+1}", f"=M{total_row}", size=10, h="center", v="center", num_fmt=MONEY_FMT, fill=FILL_TOTAL_YELLOW)

        set_cell(ws, f"J{footer_start+2}", "ИТОГО, долг + неустойка, руб.", bold=True, size=10, h="left", v="center")
        set_cell(ws, f"M{footer_start+2}", f"=M{footer_start}+M{footer_start+1}", bold=True, size=10, h="center", v="center", num_fmt=MONEY_FMT, fill=FILL_TOTAL_LIME)

        from openpyxl.styles import Border, Side
        thin = Side(style="thin", color="FF000000")
        ws[f"M{footer_start+2}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        if include_sprav:
            sprav_start = footer_start + 5

            set_cell(ws, f"J{sprav_start}", "СПРАВОЧНО", bold=True, size=13, h="left", v="center")

            thin = Side(style="thin", color="FF000000")
            for col in ("J", "K", "L", "M"):
                c = ws[f"{col}{sprav_start}"]
                c.border = (c.border.copy(bottom=thin) if c.border else Border(bottom=thin))

            set_cell(ws, f"J{sprav_start+1}", "Цена иска, руб.", size=13, h="left", v="center")
            set_cell(
                ws,
                f"M{sprav_start+1}",
                f"=M{sprav_start+2}+M{sprav_start+3}",
                bold=True,
                size=13,
                h="right",
                v="center",
                num_fmt=MONEY_FMT,
                fill=FILL_TOTAL_BLUE,
            )

            set_cell(ws, f"J{sprav_start+2}", "в т.ч. основной долг", size=13, h="left", v="center")
            set_cell(ws, f"M{sprav_start+2}", f"=SUM(M{footer_start})", size=13, h="right", v="center", num_fmt=MONEY_FMT)

            set_cell(ws, f"J{sprav_start+3}", "в т.ч. неустойка", size=13, h="left", v="center")
            set_cell(ws, f"M{sprav_start+3}", f"=SUM(M{footer_start+1})", size=13, h="right", v="center", num_fmt=MONEY_FMT)

            r = sprav_start + 4
        else:
            sprav_start = None

    last_row = max(ws.max_row, r - 1)
    return RenderTail(last_row=last_row, total_row=total_row, footer_start_row=footer_start, sprav_start_row=sprav_start)


# =============================================================================
# Header render (A2:M14)
# =============================================================================

def _render_header(ws: Worksheet, body: Any) -> None:
    debtor_name = (body.debtor.name or "").upper()
    contract_no = body.contract.number
    contract_date = body.contract.date
    category = body.category or ""
    calc_date = body.calc_date

    # A2 title (NO merge in model XLSX)
    set_cell(
        ws,
        "A2",
        "Расчет долга и неустойки в связи с просрочкой в погашении задолженности за поставленные энергоресурсы",
        bold=True,
        size=16,
        h="left",
        v="center",
    )

    # Row 4: debtor
    set_cell(ws, "A4", "Должник:", bold=True, size=13, h="left", v="center")
    ws.merge_cells("B4:M4")
    set_cell(ws, "B4", debtor_name, bold=True, size=13, h="left", v="center")

    # Row 6: contract
    set_cell(ws, "A6", "Основание:", bold=True, size=13, h="left", v="center")
    ws.merge_cells("B6:M6")
    set_cell(
        ws,
        "B6",
        f"Договор № {contract_no} от {contract_date}",
        bold=True,
        size=13,
        h="left",
        v="center",
        fill=FILL_GRAY,
    )

    # Row 8: category (B8:C8 merged, thick outer border)
    set_cell(ws, "A8", "Категория:", bold=True, size=13, h="left", v="center")
    ws.merge_cells("B8:C8")

    DISPLAY_CATEGORY = {
    "Управляющая организация": "УК",
    "ТСЖ, ЖСК, ЖК": "ТСЖ, ЖСК, ЖК",
    "Прочие": "Прочие",
    "Собственники жилых помещений в МКД": "Собственники жилых помещений в МКД",
    "Собственники нежилых помещений в МКД": "Собственники нежилых помещений в МКД",
    }

    category_display = DISPLAY_CATEGORY.get(category, category)

    set_cell(ws, "B8", category_display, bold=True, size=13, h="center", v="center")

    apply_thick_outer_border(ws, "B8", "C8")

    # Row 10: calc date (B10:C10 merged, thick outer border)
    set_cell(ws, "A10", "На дату:", bold=True, size=13, h="left", v="center")
    ws.merge_cells("B10:C10")
    dt_calc = _parse_date_ddmmyyyy(calc_date)
    set_cell(ws, "B10", dt_calc, bold=True, size=13, h="center", v="center", num_fmt=DATE_FMT)
    apply_thick_outer_border(ws, "B10", "C10")

    # Rows 12–13: column headers
    headers_af = [
        ("A", "Период"),
        ("B", "Примечание"),
        ("C", "Начислено"),
        ("D", "Оплачено"),
        ("E", "Дата"),
        ("F", "Долг"),
    ]
    for col, title in headers_af:
        ws.merge_cells(f"{col}12:{col}13")
        set_cell(ws, f"{col}12", title, bold=True, size=11, h="center", v="center", fill=FILL_YELLOW)

    # G–I grouped header
    ws.merge_cells("G12:I12")
    set_cell(ws, "G12", "Период просрочки", bold=True, size=11, h="left", v="center", fill=FILL_YELLOW)
    set_cell(ws, "G13", "с", size=11, h="center", v="center", fill=FILL_YELLOW_LIGHT)
    set_cell(ws, "H13", "по", size=11, h="center", v="center", fill=FILL_YELLOW_LIGHT)
    set_cell(ws, "I13", "дней", size=11, h="center", v="center", fill=FILL_YELLOW_LIGHT)

    # J–M merged vertically, yellow fill
    headers_jm = [("J", "Ставка"), ("K", "Доля ставки*"), ("L", "Формула"), ("M", "Неустойка")]
    for col, title in headers_jm:
        ws.merge_cells(f"{col}12:{col}13")
        set_cell(ws, f"{col}12", title, bold=True, size=11, h="center", v="center", fill=FILL_YELLOW)

    # thin outer borders for header blocks
    for col in "ABCDEF":
        apply_thin_outer_border(ws, f"{col}12", f"{col}13")
    apply_thin_outer_border(ws, "G12", "I12")
    for col in "GHI":
        apply_thin_outer_border(ws, f"{col}13", f"{col}13")
    for col in "JKLM":
        apply_thin_outer_border(ws, f"{col}12", f"{col}13")

    # Row 14: column numbers 1..13, gray fill + thin border
    for idx, col in enumerate(list("ABCDEFGHIJKLM"), start=1):
        set_cell(ws, f"{col}14", idx, size=11, h="center", v="center", fill=FILL_GRAY)
        apply_thin_outer_border(ws, f"{col}14", f"{col}14")


# =============================================================================
# Low-level parsing helpers
# =============================================================================

def _parse_date_ddmmyyyy(s: str) -> datetime:
    return datetime.strptime(s, "%d.%m.%Y")


def _dec_money(s: str) -> Decimal:
    return Decimal(s).quantize(Decimal("0.01"))
