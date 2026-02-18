from __future__ import annotations

from openpyxl.styles import PatternFill, Side

# ---------------------------------------------------------------------
# Base typography / number formats
# ---------------------------------------------------------------------

FONT_NAME = "Arial"

# Excel date format
DATE_FMT = r"dd\.mm\.yyyy"
MONEY_FMT = "#,##0.00"

# ---------------------------------------------------------------------
# Column widths (keep exactly as used by renderer/style_apply)
# ---------------------------------------------------------------------

# ---------------------------------------------------------------------
# Column widths (exact, to match the model XLSX 1:1)
# ---------------------------------------------------------------------

COLUMN_WIDTHS = {
    "A": 20.0,
    "B": 13.0,
    "C": 18.0,
    "D": 13.0,
    "E": 15.0,
    "F": 18.0,
    "G": 10.0,
    "H": 13.0,
    "I": 13.0,
    "J": 13.0,
    "K": 15.0,
    "L": 13.0,
    "M": 18.0,
}

# ---------------------------------------------------------------------
# Row heights (explicit only where required)
# ---------------------------------------------------------------------

ROW_HEIGHTS = {
    2: 20.65,
    4: 16.9,
    6: 16.9,
    8: 16.9,
    10: 16.9,
    # Footer / spravochno block
    72: 16.9,
    73: 16.9,
    74: 16.5,
    75: 16.5,
}

# ---------------------------------------------------------------------
# Fills (ARGB строго!)
# ---------------------------------------------------------------------

FILL_YELLOW = PatternFill("solid", fgColor="FFFFFF00")
FILL_YELLOW_LIGHT = PatternFill("solid", fgColor="FFFFFF99")
FILL_GRAY = PatternFill("solid", fgColor="FFD9D9D9")
FILL_GREEN = PatternFill("solid", fgColor="FFDDE8CB")

# --- Totals / footer fills (as in model XLSX) ---
FILL_TOTAL_PEACH = PatternFill("solid", fgColor="FFFFDAB9")   # ИТОГО
FILL_TOTAL_LIME  = PatternFill("solid", fgColor="FFD4EA6B")   # ИТОГО долг+неустойка
FILL_TOTAL_BLUE  = PatternFill("solid", fgColor="FFDDEBF7")   # Цена иска
FILL_TOTAL_YELLOW = PatternFill("solid", fgColor="FFFFFFD7")  # Всего неустойка (светло-жёлтая)

# ---------------------------------------------------------------------
# Borders
# ---------------------------------------------------------------------

THIN = Side(style="thin", color="FF000000")
THICK = Side(style="thick", color="FF000000")
