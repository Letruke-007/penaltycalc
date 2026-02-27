from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import Workbook

# FIX: runtime package root is "app", not "backend"
from app.contracts.statement import Statement
from app.excel.renderer import render_statement_sheet, render_statements_sheet


def _validate_ddmmyyyy(value: str) -> str:
    try:
        datetime.strptime(value, "%d.%m.%Y")
    except ValueError as e:
        raise SystemExit(
            f"--calc-date must be in DD.MM.YYYY format, got: {value}"
        ) from e
    return value


def _apply_overrides(
    raw: Dict[str, Any],
    *,
    calc_date: Optional[str] = None,
    category: Optional[str] = None,
    overdue_start_day: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Apply orchestration-level overrides over raw Statement JSON dict.

    IMPORTANT:
      - Does NOT mutate input dict (returns a shallow-copied updated dict).
      - Does NOT write anything back to disk.
    """
    out = dict(raw)
    st = dict(out.get("statement") or {})
    out["statement"] = st

    if calc_date is not None:
        st["calc_date"] = calc_date

    if category is not None:
        # Must match one of the exact category strings used in penalty_rules.py
        st["category"] = category

    if overdue_start_day is not None:
        # User-selected day-of-month (1..31) used to compute overdue_start for each period.
        st["overdue_start_day"] = int(overdue_start_day)

    return out


def build_xlsx_from_statement_json(
    in_json_path: Path,
    out_xlsx_path: Path,
    *,
    calc_date_override: Optional[str] = None,
    category_override: Optional[str] = None,
    overdue_start_day_override: Optional[int] = None,
    add_state_duty: bool = False,
) -> None:
    """
    JSON (Statement v1.1/v1.2) -> XLSX.

    IMPORTANT:
      - JSON on disk is NOT modified.
      - Rendering is delegated entirely to app/excel/renderer.py
      - Overrides are applied in-memory at orchestration level.
    """
    raw = json.loads(in_json_path.read_text(encoding="utf-8"))

    # Accept older JSON produced by earlier parser runs (schema_version 1.0),
    # but do NOT modify the JSON file on disk.
    sv = raw.get("schema_version")
    if sv == "1.0":
        raw = dict(raw)
        raw["schema_version"] = "1.1"

    # Apply orchestration overrides (in-memory only)
    raw = _apply_overrides(
        raw,
        calc_date=calc_date_override,
        category=category_override,
        overdue_start_day=overdue_start_day_override,
    )

    # Validate against current Statement contract
    stmt = Statement.model_validate(raw)

    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    # All rendering (header + table) is done in renderer.py
    render_statement_sheet(ws, stmt, add_state_duty=add_state_duty)

    out_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx_path)


# --------------------------------------------------------------------
# Public API expected by ProcessingService: json_to_xlsx(json_path, xlsx_path)
# --------------------------------------------------------------------
def json_to_xlsx(
    json_path: Path, xlsx_path: Path, *, add_state_duty: bool = False
) -> None:
    stmt = Statement.model_validate_json(Path(json_path).read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    render_statement_sheet(ws, stmt, add_state_duty=add_state_duty)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def build_xlsx_from_many_statement_jsons(
    in_json_paths: list[Path],
    out_xlsx_path: Path,
    *,
    add_state_duty: bool = False,
) -> None:
    """Build ONE XLSX from multiple Statement JSON files.

    Used for merged output (one debtor, several contracts).
    """
    stmts: list[Statement] = []
    for p in in_json_paths:
        raw = json.loads(Path(p).read_text(encoding="utf-8"))
        sv = raw.get("schema_version")
        if sv == "1.0":
            raw = dict(raw)
            raw["schema_version"] = "1.1"
        stmts.append(Statement.model_validate(raw))

    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    render_statements_sheet(ws, stmts, add_state_duty=add_state_duty)

    out_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx_path)


def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Build XLSX from Statement JSON")
    p.add_argument(
        "in_json",
        nargs="?",
        default="backend/test_files/json/07.620535-ТЭ  03.2025-04.2025.json",
        help="Path to input Statement JSON",
    )
    p.add_argument(
        "out_xlsx",
        nargs="?",
        default="out/result.xlsx",
        help="Path to output XLSX",
    )

    p.add_argument(
        "--calc-date",
        dest="calc_date",
        type=_validate_ddmmyyyy,
        default=None,
        help="Override statement.calc_date (DD.MM.YYYY) over JSON (drives B10 and penalty horizon).",
    )
    p.add_argument(
        "--category",
        dest="category",
        default=None,
        help='Override statement.category over JSON (e.g. "Управляющая организация").',
    )
    p.add_argument(
        "--overdue-start-day",
        dest="overdue_start_day",
        type=int,
        default=None,
        help="Override statement.overdue_start_day (1..31) over JSON (drives overdue start in next month).",
    )
    p.add_argument(
        "--add-state-duty",
        dest="add_state_duty",
        action="store_true",
        help="Add state duty (госпошлина) row in output XLSX (derived from claim price).",
    )

    return p


if __name__ == "__main__":
    args = _build_arg_parser().parse_args()

    in_json = Path(args.in_json)
    out_xlsx = Path(args.out_xlsx)

    print(f"[TEST] Build XLSX from {in_json}")
    build_xlsx_from_statement_json(
        in_json,
        out_xlsx,
        calc_date_override=args.calc_date,
        category_override=args.category,
        overdue_start_day_override=args.overdue_start_day,
        add_state_duty=bool(args.add_state_duty),
    )
    print(f"[OK] XLSX written to {out_xlsx.resolve()}")
